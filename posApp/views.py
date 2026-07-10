from pickle import FALSE
from django.shortcuts import redirect, render
from django.http import HttpResponse
from django.http import JsonResponse
from posApp.models import Category, Products, Sales, salesItems, Restock
from django.utils.safestring import mark_safe
from django.db.models import Count, Sum
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
import json, sys
from datetime import date, datetime
from django.db.models import OuterRef, Subquery
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse


def parse_number(value, default=0, as_int=False):
    if value is None:
        return default
    value = str(value).strip()
    if value == "":
        return default
    # Remove currency labels and whitespace
    value = value.replace("Rp", "").replace("rp", "").replace(" ", "")
    # Normalize Indonesian thousand/decimal separators
    if "," in value and "." in value:
        value = value.replace(".", "").replace(",", ".")
    elif "." in value and "," not in value:
        parts = value.split(".")
        if len(parts) > 1 and all(len(part) == 3 for part in parts[1:]):
            value = "".join(parts)
    elif "," in value and "." not in value:
        parts = value.split(",")
        if len(parts) > 1 and all(len(part) == 3 for part in parts[1:]):
            value = "".join(parts)
        else:
            value = value.replace(",", ".")
    try:
        result = float(value)
        return int(result) if as_int else result
    except ValueError:
        return default

# Login
def login_user(request):
    logout(request)
    resp = {"status":'failed','msg':''}
    username = ''
    password = ''
    if request.POST:
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                resp['status']='success'
            else:
                resp['msg'] = "Username atau Password salah!"
        else:
            resp['msg'] = "Username atau Password salah!"
    return HttpResponse(json.dumps(resp),content_type='application/json')

#Logout
def logoutuser(request):
    logout(request)
    return redirect('/')

# Create your views here.
@login_required
def home(request):
    now = datetime.now()
    current_year = now.strftime("%Y")
    current_month = now.strftime("%m")
    current_day = now.strftime("%d")
    categories = len(Category.objects.all())
    products = len(Products.objects.all())
    transaction = len(Sales.objects.filter(
        date_added__year=current_year,
        date_added__month = current_month,
        date_added__day = current_day
    ))
    today_sales = Sales.objects.filter(
        date_added__year=current_year,
        date_added__month = current_month,
        date_added__day = current_day
    ).all()
    total_sales = sum(today_sales.values_list('grand_total',flat=True))
    top_products = (
        salesItems.objects.values('product_id', 'product_id__name')
        .annotate(total_qty=Sum('qty'))
        .order_by('-total_qty')[:10]
    )
    context = {
        'page_title':'Home',
        'categories' : categories,
        'products' : products,
        'transaction' : transaction,
        'total_sales' : total_sales,
        'top_products': top_products,
    }
    return render(request, 'posApp/home.html',context)


def about(request):
    context = {
        'page_title':'About',
    }
    return render(request, 'posApp/about.html',context)

#Categories
@login_required
def category(request):
    category_list = Category.objects.all()
    context = {
        'page_title':'Category List',
        'category':category_list,
    }
    return render(request, 'posApp/category.html',context)
@login_required
def manage_category(request):
    category = {}
    if request.method == 'GET':
        data =  request.GET
        id = ''
        if 'id' in data:
            id= data['id']
        if id.isnumeric() and int(id) > 0:
            category = Category.objects.filter(id=id).first()
    
    context = {
        'category' : category
    }
    return render(request, 'posApp/manage_category.html',context)

@login_required
def save_category(request):
    data =  request.POST
    resp = {'status':'failed'}
    try:
        if (data['id']).isnumeric() and int(data['id']) > 0 :
            save_category = Category.objects.filter(id = data['id']).update(name=data['name'], description = data['description'])
        else:
            save_category = Category(name=data['name'], description = data['description'])
            save_category.save()
        resp['status'] = 'success'
        messages.success(request, 'Kategori berhasil disimpan.')
    except:
        resp['status'] = 'failed'
    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def delete_category(request):
    data =  request.POST
    resp = {'status':''}
    try:
        Category.objects.filter(id = data['id']).delete()
        resp['status'] = 'success'
        messages.success(request, 'Category Successfully deleted.')
    except:
        resp['status'] = 'failed'
    return HttpResponse(json.dumps(resp), content_type="application/json")

# Products
@login_required
def products(request):
    latest_restock = Restock.objects.filter(
        product_id=OuterRef('pk')
    ).order_by('-restock_date')

    product_list = Products.objects.annotate(
        latest_f_restock=Subquery(latest_restock.values('f_restock')[:1]),
        latest_sale=Subquery(latest_restock.values('sale')[:1]),
        latest_restock_date=Subquery(latest_restock.values('restock_date')[:1])
    )
    context = {
        'page_title':'Product List',
        'products':product_list,
    }
    return render(request, 'posApp/products.html',context)

def restock_list(request):
    # Ambil data restock terakhir per produk
    latest_restock = Restock.objects.filter(
        product_id=OuterRef('pk')
    ).order_by('-restock_date')

    # Ambil produk dengan stok kurang dari 5
    restock_products = Products.objects.filter(stock__lt=5).annotate(
        latest_f_restock=Subquery(latest_restock.values('f_restock')[:1]),
        latest_sale=Subquery(latest_restock.values('sale')[:1]),
        latest_restock_date=Subquery(latest_restock.values('restock_date')[:1])
    ).order_by('stock')

    context = {
        'page_title': 'Daftar produk yang perlu di restok',
        'products': restock_products,
    }

    return render(request, 'posApp/restock_list.html', context)

@login_required
def manage_products(request):
    product = {}
    categories = Category.objects.all()

    if request.method == 'GET':
        data = request.GET
        id = data.get('id', '')

        if id.isnumeric() and int(id) > 0:
            product = Products.objects.filter(id=id).first()

    # --- Dapatkan kode terakhir ---
    last_product = Products.objects.order_by('-id').first()
    if last_product:
        # Ambil angka dari kode terakhir (contoh: 005 -> 5)
        last_code_num = int(''.join(filter(str.isdigit, last_product.code)))
        next_code_num = last_code_num + 1
    else:
        next_code_num = 1

    # Format kode baru misal: P001, P002, ...
    next_code = f"{next_code_num:03d}"

    context = {
        'product': product,
        'categories': categories,
        'next_code': next_code
    }
    return render(request, 'posApp/manage_product.html', context)

def test(request):
    categories = Category.objects.all()
    context = {
        'categories' : categories
    }
    return render(request, 'posApp/test.html',context)

@login_required
def save_product(request):
    data =  request.POST
    resp = {'status':'failed'}
    id= ''
    if 'id' in data:
        id = data['id']
    if id.isnumeric() and int(id) > 0:
        check = Products.objects.exclude(id=id).filter(code=data['code']).all()
    else:
        check = Products.objects.filter(code=data['code']).all()
    if len(check) > 0 :
        resp['msg'] = "Product Code Already Exists in the database"
    else:
        category = Category.objects.filter(id = data['category_id']).first()
        
        try:
            # Update
            if (data['id']).isnumeric() and int(data['id']) > 0 :
                add_stock = parse_number(data['stock'], default=0, as_int=True)
                new_stock = Products.objects.filter(id = data['id']).first().stock + add_stock
                price = parse_number(data['price'], default=0.0)
                modal = parse_number(data['modal'], default=0.0)
                if modal > price:
                    resp['msg'] = 'Modal tidak boleh lebih besar dari harga jual'
                    return HttpResponse(json.dumps(resp), content_type="application/json")
                Products.objects.filter(id = data['id']).update(code=data['code'], category_id=category, name=data['name'], price = price, modal=modal, stock=new_stock)
                restock = Restock.objects.filter(product_id=data['id']).last()
                Restock.objects.filter(id=restock.pk).update(sale=0, f_restock_before=restock.f_restock, f_restock=0)
            # create   
            else:
                stock = parse_number(data['stock'], default=0, as_int=True)
                price = parse_number(data['price'], default=0.0)
                modal = parse_number(data['modal'], default=0.0)
                if modal > price:
                    resp['msg'] = 'Modal tidak boleh lebih besar dari harga jual'
                    return HttpResponse(json.dumps(resp), content_type="application/json")
                status = True if stock > 0 else False
                save_product = Products(code=data['code'], category_id=category, name=data['name'], price = price, modal=modal, status = status, stock=stock)
                save_product.save()
                Restock(product_id=save_product, f_restock_before=stock).save()

            resp['status'] = 'success'
            messages.success(request, 'Product Successfully saved.')
        except:
            resp['status'] = 'failed'
    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def delete_product(request):
    data =  request.POST
    resp = {'status':''}
    try:
        Products.objects.filter(id = data['id']).delete()
        resp['status'] = 'success'
        messages.success(request, 'Product Successfully deleted.')
    except:
        resp['status'] = 'failed'
    return HttpResponse(json.dumps(resp), content_type="application/json")

@login_required
def pos(request):
    products = Products.objects.filter(status = 1)
    product_json = []
    for product in products:
        product_json.append({
            'id': product.id,
            'name': product.name,
            'price': float(product.price),
            'modal': float(product.modal),
            'stock': int(product.stock)
        })
    context = {
        'page_title' : "Point of Sale",
        'products' : products,
        'product_json' : json.dumps(product_json)
    }
    # return HttpResponse('')
    return render(request, 'posApp/pos.html',context)

@login_required
def checkout_modal(request):
    grand_total = 0
    if 'grand_total' in request.GET:
        grand_total = request.GET['grand_total']
    context = {
        'grand_total' : grand_total,
    }
    return render(request, 'posApp/checkout.html', context)

def predict_f_restock(id):
    old_restock = Restock.objects.filter(pk=id).first()
    new_f_restock = (0.1 * old_restock.sale) + (1 - 0.1) * old_restock.f_restock_before
    Restock.objects.filter(pk=id).update(f_restock=new_f_restock)

@login_required
def save_pos(request):
    resp = {'status':'failed','msg':''}
    data = request.POST
    pref = datetime.now().year + datetime.now().year
    i = 1
    while True:
        code = '{:0>5}'.format(i)
        i += int(1)
        check = Sales.objects.filter(code = str(pref) + str(code)).all()
        if len(check) <= 0:
            break
    code = str(pref) + str(code)

    try:
        sub_total = parse_number(data.get('sub_total', ''), default=0.0)
        grand_total = parse_number(data.get('grand_total', ''), default=0.0)
        tendered_amount = parse_number(data.get('tendered_amount', ''), default=0.0)
        amount_change = parse_number(data.get('amount_change', ''), default=0.0)
        if amount_change < 0:
            resp['msg'] = 'Jumlah kembalian tidak valid.'
            return HttpResponse(json.dumps(resp), content_type="application/json")

        sale_obj = Sales(
            code=code,
            sub_total=sub_total,
            grand_total=grand_total,
            tendered_amount=tendered_amount,
            amount_change=amount_change,
            profit=0.0,
        )
        sale_obj.save()
        sale_id = sale_obj.pk
        sale_profit = 0
        i = 0
        for prod in data.getlist('product_id[]'):
            product_id = prod
            sale = Sales.objects.filter(id=sale_id).first()
            product = Products.objects.filter(id=product_id).first()
            latest_restock = Restock.objects.filter(product_id=product_id).last()
            qty = parse_number(data.getlist('qty[]')[i], default=0, as_int=True)
            terjual = latest_restock.sale
            price = parse_number(data.getlist('price[]')[i], default=0.0)
            modal = parse_number(data.getlist('modal[]')[i], default=0.0)
            if modal > price:
                resp['msg'] = 'Modal tidak boleh lebih besar dari harga jual.'
                return HttpResponse(json.dumps(resp), content_type="application/json")
            total = qty * price
            profit = (price - modal) * qty
            sale_profit += profit
            stock_now = product.stock - qty
            sale_now = terjual + qty
            Products.objects.filter(id=product_id).update(stock=stock_now)
            Restock.objects.filter(id=latest_restock.pk).update(sale=sale_now)
            predict_f_restock(latest_restock.pk)
            salesItems(
                sale_id=sale,
                product_id=product,
                qty=qty,
                price=price,
                modal=modal,
                total=total,
                profit=profit,
            ).save()
            i += 1
        Sales.objects.filter(id=sale_id).update(profit=sale_profit)
            
        resp['status'] = 'success'
        resp['sale_id'] = sale_id
        messages.success(request, "Sale Record has been saved.")

        # 🟦 Tambahan: cek stok rendah dan tampilkan alert
        low_stock_products = Products.objects.filter(stock__lt=5)
        if low_stock_products.exists():
            low_list = ", ".join([p.name for p in low_stock_products])
            alert_message = (
                f"⚠️ Stok menipis untuk produk: {low_list}.<br>"
                f"<a href='/restock-list/' class='btn btn-sm btn-primary mt-2'>"
                f"Lihat Daftar Restok</a>"
            )
            messages.warning(request, mark_safe(alert_message))  # ✅ aman untuk HTML dalam messages

    except Exception as e:
        resp['msg'] = "An error occured"
        print("Unexpected error:", sys.exc_info()[0])
        resp['msg'] = f"An error occurred: {str(e)}"
        print("Unexpected error:", e)
        traceback.print_exc()
    return HttpResponse(json.dumps(resp),content_type="application/json")


@login_required
def salesList(request):
    sales = Sales.objects.all().order_by('-date_added')
    month = request.GET.get('month', '')
    start_month = request.GET.get('start_month', '')
    end_month = request.GET.get('end_month', '')
    year = request.GET.get('year', '')

    if year:
        sales = sales.filter(date_added__year=year)
    if month:
        sales = sales.filter(date_added__month=month)
    elif start_month and end_month:
        sales = sales.filter(date_added__month__gte=start_month, date_added__month__lte=end_month)

    sale_data = []
    for sale in sales:
        data = {}
        for field in sale._meta.get_fields(include_parents=False):
            if field.related_model is None:
                data[field.name] = getattr(sale,field.name)
        data['items'] = salesItems.objects.filter(sale_id=sale).all()
        data['item_count'] = len(data['items'])
        if data['item_count'] <= 0:
            continue
        if 'tax_amount' in data:
            data['tax_amount'] = format(float(data['tax_amount']),'.2f')
        sale_data.append(data)

    year_choices = list(Sales.objects.dates('date_added', 'year', order='DESC'))
    year_choices = [y.year for y in year_choices]
    month_choices = [
        (1, 'Januari'), (2, 'Februari'), (3, 'Maret'), (4, 'April'),
        (5, 'Mei'), (6, 'Juni'), (7, 'Juli'), (8, 'Agustus'),
        (9, 'September'), (10, 'Oktober'), (11, 'November'), (12, 'Desember')
    ]

    context = {
        'page_title':'Sales Transactions',
        'sale_data':sale_data,
        'selected_month': month,
        'selected_start_month': start_month,
        'selected_end_month': end_month,
        'selected_year': year,
        'year_choices': year_choices,
        'month_choices': month_choices,
    }
    return render(request, 'posApp/sales.html', context)

@login_required
def export_sales_excel(request):
    month = request.GET.get('month', '')
    start_month = request.GET.get('start_month', '')
    end_month = request.GET.get('end_month', '')
    year = request.GET.get('year', '')

    sales = Sales.objects.all().order_by('-date_added')
    if year:
        sales = sales.filter(date_added__year=year)
    if month:
        sales = sales.filter(date_added__month=month)
    elif start_month and end_month:
        sales = sales.filter(date_added__month__gte=start_month, date_added__month__lte=end_month)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Data Transaksi'

    headers = ['Tanggal', 'Kode Transaksi', 'Subtotal', 'Grand Total', 'Dibayar', 'Kembalian', 'Laba']
    ws.append(headers)

    for sale in sales:
        ws.append([
            sale.date_added.strftime('%Y-%m-%d %H:%M'),
            sale.code,
            sale.sub_total,
            sale.grand_total,
            sale.tendered_amount,
            sale.amount_change,
            sale.profit,
        ])

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal='center')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="laporan_transaksi.xlsx"'
    wb.save(response)
    return response

@login_required
def profitList(request):
    profits = Sales.objects.all().order_by('-date_added')
    year = request.GET.get('year', '')
    month = request.GET.get('month', '')
    start_month = request.GET.get('start_month', '')
    end_month = request.GET.get('end_month', '')
    date = request.GET.get('date', '')

    if year:
        profits = profits.filter(date_added__year=year)
    if date:
        profits = profits.filter(date_added__date=date)
    if month:
        profits = profits.filter(date_added__month=month)
    elif start_month and end_month:
        profits = profits.filter(date_added__month__gte=start_month, date_added__month__lte=end_month)

    profit_data = []
    total_profit = 0
    for sale in profits:
        data = {
            'id': sale.id,
            'date_added': sale.date_added,
            'profit': sale.profit,
        }
        total_profit += sale.profit
        profit_data.append(data)

    year_choices = list(Sales.objects.dates('date_added', 'year', order='DESC'))
    year_choices = [y.year for y in year_choices]
    month_choices = [
        (1, 'Januari'), (2, 'Februari'), (3, 'Maret'), (4, 'April'),
        (5, 'Mei'), (6, 'Juni'), (7, 'Juli'), (8, 'Agustus'),
        (9, 'September'), (10, 'Oktober'), (11, 'November'), (12, 'Desember')
    ]

    context = {
        'page_title': 'Laporan Laba',
        'profit_data': profit_data,
        'total_profit': total_profit,
        'selected_year': year,
        'selected_month': month,
        'selected_start_month': start_month,
        'selected_end_month': end_month,
        'selected_date': date,
        'year_choices': year_choices,
        'month_choices': month_choices,
    }
    return render(request, 'posApp/profit.html', context)

@login_required
def export_profit_excel(request):
    year = request.GET.get('year', '')
    month = request.GET.get('month', '')
    start_month = request.GET.get('start_month', '')
    end_month = request.GET.get('end_month', '')
    date = request.GET.get('date', '')

    profits = Sales.objects.all().order_by('-date_added')
    if year:
        profits = profits.filter(date_added__year=year)
    if date:
        profits = profits.filter(date_added__date=date)
    if month:
        profits = profits.filter(date_added__month=month)
    elif start_month and end_month:
        profits = profits.filter(date_added__month__gte=start_month, date_added__month__lte=end_month)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Laporan Laba'

    headers = ['Tanggal', 'Kode Transaksi', 'Jumlah Item', 'Total Penjualan', 'Laba']
    ws.append(headers)

    for sale in profits:
        ws.append([
            sale.date_added.strftime('%Y-%m-%d %H:%M'),
            sale.code,
            salesItems.objects.filter(sale_id=sale).count(),
            sale.grand_total,
            sale.profit,
        ])

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal='center')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="laporan_laba.xlsx"'
    wb.save(response)
    return response

@login_required
def receipt(request):
    id = request.GET.get('id')
    sales = Sales.objects.filter(id = id).first()
    transaction = {}
    for field in Sales._meta.get_fields():
        if field.related_model is None:
            transaction[field.name] = getattr(sales,field.name)
    if 'tax_amount' in transaction:
        transaction['tax_amount'] = format(float(transaction['tax_amount']))
    ItemList = salesItems.objects.filter(sale_id = sales).all()
    context = {
        "transaction" : transaction,
        "salesItems" : ItemList
    }

    return render(request, 'posApp/receipt.html',context)
    # return HttpResponse('')

@login_required
def delete_sale(request):
    resp = {'status':'failed', 'msg':''}
    id = request.POST.get('id')
    try:
        delete = Sales.objects.filter(id = id).delete()
        resp['status'] = 'success'
        messages.success(request, 'Sale Record has been deleted.')
    except Exception as e:
        resp['msg'] = "An error occured"
        print("Unexpected error:", sys.exc_info()[0])
        print(e)
    return HttpResponse(json.dumps(resp), content_type='application/json')

@login_required
def delete_all_sales(request):
    resp = {'status':'failed', 'msg':''}
    try:
        Sales.objects.all().delete()
        resp['status'] = 'success'
        messages.success(request, 'Semua riwayat transaksi telah dihapus.')
    except Exception as e:
        resp['msg'] = "An error occured"
        print("Unexpected error:", sys.exc_info()[0])
        print(e)
    return HttpResponse(json.dumps(resp), content_type='application/json')

def guest_product_display(request):
    # Ambil semua produk
    products = Products.objects.all().order_by('name')
    context = {
        'page_title': 'Daftar Produk Koperasi',
        'products': products
    }
    return render(request, 'posApp/guest_display.html', context)